import os
import xlrd
import openpyxl

# 定义一个函数，用于将.xls文件转换为.xlsx文件
def convert_xls_to_xlsx(folder_path):
    for file_name in os.listdir(folder_path):
        # 判断文件是否为.xls文件
        if file_name.endswith('.xls'):
            file_path = os.path.join(folder_path, file_name)
            # 使用xlrd库打开.xls文件，并读取数据
            workbook = xlrd.open_workbook(file_path)
            sheets = workbook.sheet_names()
            data = {}
            for sheet_name in sheets:
                sheet = workbook.sheet_by_name(sheet_name)
                rows = []
                for row_idx in range(sheet.nrows):
                    row = []
                    for col_idx in range(sheet.ncols):
                        cell_value = sheet.cell_value(row_idx, col_idx)
                        row.append(cell_value)
                    rows.append(row)
                data[sheet_name] = rows
            # 使用openpyxl库将数据写入新的.xlsx文件
            new_file_name = file_name.replace('.xls', '.xlsx')
            new_file_path = os.path.join(folder_path, new_file_name)
            workbook = openpyxl.Workbook()
            for sheet_name, rows in data.items():
                sheet = workbook.create_sheet(sheet_name)
                for row_idx, row in enumerate(rows):
                    for col_idx, cell_value in enumerate(row):
                        sheet.cell(row=row_idx+1, column=col_idx+1, value=cell_value)
            workbook.save(new_file_path)
    print("转换完毕:)")

# 定义一个函数，用于删除指定路径下的所有后缀为.xls的文件
def delete_xls_files(folder_path):
    for filename in os.listdir(folder_path):
        if filename.endswith('.xls'):
            os.remove(os.path.join(folder_path, filename))
            print(f'{filename} has been deleted.')
    print("删除完毕:)")

# 定义一个函数，用于将多个Excel文件中的数据整合到一个新的Excel文件中
def merge_excel_files(folder_path, output_file_path):
    workbook = openpyxl.Workbook()
    workbook.remove(workbook.active)
    for file_name in os.listdir(folder_path):
        # 判断文件是否为Excel文件
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(folder_path, file_name)
            # 使用openpyxl库读取Excel文件中的数据，并写入新的Excel文件中
            file_name_without_ext = os.path.splitext(file_name)[0]
            workbook.create_sheet(file_name_without_ext)
            excel_file = openpyxl.load_workbook(file_path)
            for sheet_name in excel_file.sheetnames:
                sheet = excel_file[sheet_name]
                new_sheet = workbook[file_name_without_ext]
                for row_idx, row in enumerate(sheet.iter_rows()):
                    for col_idx, cell in enumerate(row):
                        new_sheet.cell(row=row_idx+1, column=col_idx+1, value=cell.value)
    # 保存新的Excel文件
    workbook.save(output_file_path)
    print("整合完毕:)")

# 定义一个函数，用于将该文件中所有有数据的sheet中有数据的单元格进行居中操作
def center_cells_in_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value is not None:
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center', vertical='center')
    wb.save(filename)
    print("居中完毕:)")

# 定义一个函数，用于扩展该文件中所有有数据的列宽
import openpyxl
def adjust_column_width_in_xlsx(filename):
    wb = openpyxl.load_workbook(filename)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for column in sheet.columns:
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value)) + 1
                except:
                    pass
            adjusted_width = max_length + 15
            sheet.column_dimensions[column_letter].width = adjusted_width
    wb.save(filename)
    print("列宽修改完毕:)")

if __name__ == '__main__':
    folder_path = r'D:\（程序设计）Python\2.编程题练习\3.Excel指令\文件集合'
    output_file_path = r'D:\（程序设计）Python\2.编程题练习\3.Excel指令\志愿活动 x.xx-x.xx.xlsx'

    # 调用convert_xls_to_xlsx函数将文件夹中的.xls文件转换为.xlsx文件
    convert_xls_to_xlsx(folder_path)

    # 调用delete_xls_files函数将文件夹中的.xls文件删除
    delete_xls_files(folder_path)

    # 调用merge_excel_files函数将多个Excel文件中的数据整合到一个新的Excel文件中
    merge_excel_files(folder_path, output_file_path)

    # 调用center_cells_in_xlsx函数将该Excel中所有有数据的sheet中有数据的单元格进行居中操作
    center_cells_in_xlsx(output_file_path)

    # 调用expand_column_width_in_xlsx函数，扩展该Excel中所有有数据的列宽
    adjust_column_width_in_xlsx(output_file_path)